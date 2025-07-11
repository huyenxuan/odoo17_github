from odoo import fields, models, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook
import time

class MrpProductionWiz(models.TransientModel):
    _name = 'mrp.product.wiz'
    _description = 'Mrp Production Wizard'

    file_import = fields.Binary(string='File import', help='File import to tab component. Only support Excel file', tracking=True, attachment=False)
    file_name = fields.Char(string='File name', tracking=True)

    def action_import(self):
        time_start = time.time()
        if not self.file_import:
            raise UserError(_('No file'))
        else:
            if not self.file_name or not self.file_name.lower().endswith(('.xls', '.xlsx')):
                raise UserError(_('Please import file excel (.xls or .xlsx)'))
                
            file_data = base64.b64decode(self.file_import)
            buffer = io.BytesIO(file_data)
            try:
                workbook = load_workbook(buffer, data_only=True)
            except:
                raise UserError(_('Không thể đọc file'))
            
            sheet = workbook.active
            lines = []

            # Lấy id bản ghi MO
            active_id = self.env.context.get('active_id')
            if not active_id:
                raise UserError(_("No MO found"))
            mo = self.env['mrp.production'].browse(active_id)

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                product_name, quantity = row
                if not product_name:
                    continue

                product = self.env['product.product'].search([
                    ('name', '=', product_name),
                ], limit=1)
                if not product:
                    raise UserError(_(f"No name products have been found: '{product_name}'"))
                
                lines.append((0, 0, {
                    'product_id': product.id,
                    'name': product.name,
                    'product_uom_qty': quantity,
                }))
            
            mo.move_raw_ids = [(5, 0, 0)] + lines # Xóa xác dòng cũ và gán cái mới
        time_end = time.time()
        duration = time_end - time_start
        print(f'\nQuá trình import mất: {duration}s\n')
